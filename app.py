import os, re, io, zipfile, datetime, tempfile, subprocess, unicodedata
from flask import Flask, render_template, request, send_file, jsonify
import openpyxl
import pytesseract
import fitz
from PIL import Image

app = Flask(__name__, template_folder='.')
app.config['MAX_CONTENT_LENGTH'] = 30 * 1024 * 1024

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
PLANT_DIR = BASE_DIR


def _normalizar(s):
    """minusculas + sin tildes, para matchear nombres de archivo de plantilla
    sin importar mayusculas/minusculas o acentos (ej. 'SINGAPUR Aéreo.docx')."""
    s = s.lower()
    s = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in s if not unicodedata.combining(c))

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/generar', methods=['POST'])
def generar():
    try:
        piqueo_f    = request.files.get('piqueo')
        reporte_f   = request.files.get('reporte')
        sanitario_f = request.files.get('sanitario')
        remito_f    = request.files.get('remito')
        shipment    = request.form.get('shipment_no', '').strip()
        tipo_via    = request.form.get('tipo_via', '').strip()
        destino     = request.form.get('destino', 'malasia').strip()

        errores = []
        if not piqueo_f:    errores.append('Falta el Piqueo (.xlsx)')
        if not reporte_f:   errores.append('Falta el Reporte DOC (.xlsx)')
        if not sanitario_f: errores.append('Falta el Sanitario Provisorio (PDF)')
        if not remito_f:    errores.append('Falta el Remito (PDF)')
        if not shipment:    errores.append('Falta el numero de Shipment')
        if not tipo_via:    errores.append('Selecciona la via (Aereo o Maritimo)')
        if errores:
            return jsonify({'ok': False, 'errores': errores}), 400

        datos_piqueo = leer_piqueo(piqueo_f)
        reporte      = leer_reporte(reporte_f, shipment)
        datos_remito = leer_remito(remito_f.read())
        datos_prov   = leer_sanitario_provisorio(sanitario_f.read())

        datos = {**datos_remito, **datos_piqueo, **datos_prov}
        datos['destino'] = destino
        # Pallets: provisorio tiene prioridad, piqueo como fallback
        if not datos.get('pallets'):
            datos['pallets'] = datos_piqueo.get('pallets_piqueo')
        # Congelado: remito tiene prioridad sobre provisorio
        if datos_remito.get('es_congelado') is not None:
            datos['es_congelado'] = datos_remito['es_congelado']
        lotes_map = datos_piqueo.get('lotes_por_producto', {})
        fecha_prod_map  = datos_piqueo.get('fecha_produccion_por_producto', {})
        fecha_faena_map = datos_piqueo.get('fecha_faena_por_producto', {})
        productos_list = datos.get('productos', [])
        for prod in productos_list:
            cod = prod.get('codigo', '')
            if cod in reporte.get('descripciones', {}):
                prod['nombre_en'] = reporte['descripciones'][cod]
            else:
                prod['nombre_en'] = buscar_nombre_en(prod.get('nombre_es', ''))
            prod['lotes'] = lotes_map.get(cod, '')
            prod['fecha_produccion_prod'] = fecha_prod_map.get(cod, '')
            prod['fecha_faena_prod']      = fecha_faena_map.get(cod, '')

        # Cruce de Contramarca (formato USA con anexo) en 3 niveles, porque el
        # OCR del sanitario provisorio viene con ruido (numeros perdidos o mal
        # leidos): 1) por cantidad de cajas cuando matchea unico; 2) por nombre
        # del producto dentro del texto de la linea, para las que no tienen
        # cajas; 3) por eliminacion, si al final queda exactamente 1 producto
        # y 1 codigo valido sin asignar. Los codigos se validan contra la
        # lista real del remito (CONTRAMARCA:C164/65/66/...) para descartar
        # lecturas OCR invalidas (ej. C185 cuando el unico codigo posible es C165).
        lineas_usa = datos_prov.get('lineas_usa', [])
        codigos_validos = set(expandir_contramarcas(datos_remito.get('contramarca') or ''))
        if codigos_validos and len(codigos_validos) < len(productos_list):
            # El campo CONTRAMARCA del remito vino truncado/incompleto (la capa de
            # texto del PDF a veces corta el campo aunque visualmente se vea completo).
            # Como las contramarcas de USA son siempre un rango consecutivo, se infiere
            # el rango completo a partir del primer numero confirmado + la cantidad de
            # productos del envio.
            numeros = sorted(set(int(re.sub(r'\D', '', c)) for c in codigos_validos))
            base = numeros[0]
            codigos_validos = set('C' + str(base + i) for i in range(len(productos_list)))
        if codigos_validos:
            lineas_usa = [l for l in lineas_usa if l['contramarca'] in codigos_validos]

        usados_cod = set()
        usados_idx = set()

        por_cajas = {}
        for l in lineas_usa:
            if l['cajas']:
                por_cajas.setdefault(l['cajas'], []).append(l)
        for i, prod in enumerate(productos_list):
            cand = por_cajas.get(str(prod.get('cajas', '')), [])
            if len(cand) == 1 and cand[0]['contramarca'] not in usados_cod:
                prod['contramarca'] = cand[0]['contramarca']
                usados_cod.add(cand[0]['contramarca'])
                usados_idx.add(i)

        for l in lineas_usa:
            if l['contramarca'] in usados_cod:
                continue
            texto_l = l.get('texto', '')
            cand_idx = [i for i, p in enumerate(productos_list)
                        if i not in usados_idx and (p.get('nombre_es') or '').upper() in texto_l]
            if len(cand_idx) == 1:
                i = cand_idx[0]
                productos_list[i]['contramarca'] = l['contramarca']
                usados_cod.add(l['contramarca'])
                usados_idx.add(i)

        restantes_idx = [i for i in range(len(productos_list)) if i not in usados_idx]
        restantes_cod = [c for c in codigos_validos if c not in usados_cod]
        if len(restantes_idx) == 1 and len(restantes_cod) == 1:
            productos_list[restantes_idx[0]]['contramarca'] = restantes_cod[0]

        PATRONES_DESTINO = {
            'malasia':     'alasia',
            'singapur':    'ingapur',
            'mexico':      'exico',
            'usawclass':   'class',
            'usaorleans':  'orleans',
        }
        patron_via = 'aereo' if tipo_via == 'aereo' else 'mar'
        patron_dest = PATRONES_DESTINO.get(destino, PATRONES_DESTINO['malasia'])
        todos_docx = [f for f in os.listdir(PLANT_DIR) if f.lower().endswith('.docx')]
        candidatos = [f for f in todos_docx if patron_dest in _normalizar(f) and patron_via in _normalizar(f)]
        if not candidatos:
            candidatos = [f for f in todos_docx if patron_dest in _normalizar(f)]
        if not candidatos:
            return jsonify({'ok': False, 'errores': [
                'Plantilla no encontrada para destino=' + destino + ' via=' + tipo_via + '. Archivos: ' + str(os.listdir(PLANT_DIR))
            ]}), 500
        plantilla = os.path.join(PLANT_DIR, candidatos[0])

        with open(plantilla, 'rb') as f:
            docx_bytes = f.read()

        resultado, alertas = generar_sanitario(docx_bytes, datos, tipo_via, destino)

        nombre_archivo = 'Sanitario_' + destino.capitalize() + '_' + tipo_via + '_' + shipment + '.docx'
        resp = send_file(
            io.BytesIO(resultado), as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        if alertas:
            resp.headers['X-Alertas'] = ' | '.join(alertas)
        return resp

    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'errores': [str(e), traceback.format_exc()]}), 500


# ── PIQUEO ───────────────────────────────────────────────────────────────────

def leer_piqueo(file):
    wb = openpyxl.load_workbook(file, data_only=True)

    # Buscar entre TODAS las hojas la que tenga headers de producto (Cod Prod/Codigo).
    # No asumir que la hoja activa (wb.active) es la correcta - puede haber hojas
    # sueltas/borrador (ej. "Hoja1") marcadas como activas por accidente.
    ws = None
    rows = None
    hdr_idx = None
    for sheet in wb.worksheets:
        rows_tmp = list(sheet.iter_rows(values_only=True))
        for i, row in enumerate(rows_tmp[:5]):
            if not row: continue
            valores = [str(v or '').strip() for v in row]
            tiene_cod   = any(v in ('Cod Prod', 'Codigo') for v in valores)
            tiene_fecha = any(v in ('Fecha F', 'Fecha P', 'Fecha Ven') for v in valores)
            if tiene_cod and tiene_fecha:
                ws, rows, hdr_idx = sheet, rows_tmp, i
                break
        if ws is not None:
            break

    if ws is None:
        # Fallback: comportamiento anterior (hoja activa) por si ninguna calza
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        hdr_idx = 0
        for i, row in enumerate(rows[:5]):
            if row and any(str(v or '').strip() in ('Cod Prod', 'Producto', 'Fecha F', 'Fecha P') for v in row):
                hdr_idx = i
                break

    hdr = rows[hdr_idx]

    def col_idx(nombres):
        for i, h in enumerate(hdr):
            if h and any(n.lower() in str(h).lower() for n in nombres):
                return i
        return None

    c_cod     = col_idx(['Cod Prod', 'Codigo'])
    c_fecha_f = col_idx(['Fecha F'])
    c_fecha_p = col_idx(['Fecha P'])
    c_fecha_v = col_idx(['Fecha Ven', 'Vencimiento'])
    c_pallet  = col_idx(['Pallet'])

    faena_min = faena_max = None
    prod_min  = prod_max  = None
    venc_min  = venc_max  = None
    pallets_set = set()
    lotes_por_cod = {}        # Cod Prod -> set de fechas 'Fecha P' (YYYYMMDD), para Nº de lotes (ej. Mexico)
    prod_fechas_por_cod = {}   # Cod Prod -> [fecha_p_min, fecha_p_max] (rango de produccion por producto, ej. Orleans)
    faena_fechas_por_cod = {}  # Cod Prod -> [fecha_f_min, fecha_f_max] (rango de faena por producto, ej. Orleans)

    for row in rows[hdr_idx + 1:]:
        if not row: continue
        cod = row[c_cod] if c_cod is not None and c_cod < len(row) else None
        if not cod: continue
        cod = str(cod).strip()

        if c_pallet is not None and c_pallet < len(row) and row[c_pallet]:
            pallets_set.add(str(row[c_pallet]))

        fecha_f = row[c_fecha_f] if c_fecha_f is not None and c_fecha_f < len(row) else None
        fecha_p = row[c_fecha_p] if c_fecha_p is not None and c_fecha_p < len(row) else None
        fecha_v = row[c_fecha_v] if c_fecha_v is not None and c_fecha_v < len(row) else None

        if isinstance(fecha_f, datetime.datetime):
            faena_min = min(faena_min, fecha_f) if faena_min else fecha_f
            faena_max = max(faena_max, fecha_f) if faena_max else fecha_f
            fmin, fmax = faena_fechas_por_cod.get(cod, (fecha_f, fecha_f))
            faena_fechas_por_cod[cod] = (min(fmin, fecha_f), max(fmax, fecha_f))
        if isinstance(fecha_p, datetime.datetime):
            prod_min = min(prod_min, fecha_p) if prod_min else fecha_p
            prod_max = max(prod_max, fecha_p) if prod_max else fecha_p
            lotes_por_cod.setdefault(cod, set()).add(fecha_p.strftime('%Y%m%d'))
            pmin, pmax = prod_fechas_por_cod.get(cod, (fecha_p, fecha_p))
            prod_fechas_por_cod[cod] = (min(pmin, fecha_p), max(pmax, fecha_p))
        if isinstance(fecha_v, datetime.datetime):
            venc_min = min(venc_min, fecha_v) if venc_min else fecha_v
            venc_max = max(venc_max, fecha_v) if venc_max else fecha_v

    def fmt_rango(d_min, d_max):
        if not d_min: return None
        s = d_min.strftime('%d/%m/%Y')
        if d_max and isinstance(d_max, datetime.datetime) and d_max != d_min:
            s += ' al ' + d_max.strftime('%d/%m/%Y')
        return s

    lotes_por_producto = {
        cod: ' - '.join(sorted(fechas)) for cod, fechas in lotes_por_cod.items()
    }
    fecha_produccion_por_producto = {
        cod: fmt_rango(pmin, pmax) for cod, (pmin, pmax) in prod_fechas_por_cod.items()
    }
    fecha_faena_por_producto = {
        cod: fmt_rango(fmin, fmax) for cod, (fmin, fmax) in faena_fechas_por_cod.items()
    }

    return {
        'fecha_faena':       fmt_rango(faena_min, faena_max),
        'fecha_produccion':  fmt_rango(prod_min,  prod_max),
        'fecha_vencimiento': fmt_rango(venc_min,  venc_max),
        'pallets_piqueo':    str(len(pallets_set)) if pallets_set else None,
        'lotes_por_producto': lotes_por_producto,
        'fecha_produccion_por_producto': fecha_produccion_por_producto,
        'fecha_faena_por_producto': fecha_faena_por_producto,
    }


# ── REPORTE DOC ──────────────────────────────────────────────────────────────

def leer_reporte(file, shipment):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return {}
    hdr = None
    hdr_idx = 0
    for i, row in enumerate(rows[:5]):
        if row and any(str(v or '').strip() in ('Shipment No', 'Shipment') for v in row):
            hdr = row; hdr_idx = i; break
    if hdr is None: hdr = rows[0]
    def col(keys):
        for i, h in enumerate(hdr):
            if h and any(k.lower() in str(h).lower() for k in keys): return i
        return None
    c_ship = col(['Shipment No', 'Shipment'])
    c_code = col(['Code'])
    c_desc = col(['Description'])
    d = {'descripciones': {}}
    ship_base = shipment.split('-')[0]
    for row in rows[hdr_idx + 1:]:
        if not row or c_ship is None: continue
        sv = str(row[c_ship] or '').strip()
        if sv != shipment and not sv.startswith(ship_base): continue
        if c_code is not None and c_desc is not None and row[c_code] and row[c_desc]:
            d['descripciones'][str(row[c_code]).strip()] = str(row[c_desc]).strip()
    return d


# ── UTILIDADES NUMÉRICAS ─────────────────────────────────────────────────────

def expandir_contramarcas(campo):
    """Expande el formato abreviado de contramarcas del remito, ej.
    'C164/65/66/67/68/69/70/71/72/73/74' -> ['C164','C165',...,'C174'].
    Sirve como lista de codigos validos para validar/corregir lo leido por OCR
    en el sanitario provisorio (que a veces confunde digitos, ej. C165->C185)."""
    if not campo:
        return []
    partes = [p.strip() for p in re.split(r'[/,]', campo) if p.strip()]
    if not partes:
        return []
    primero = re.sub(r'\D', '', partes[0])
    if not primero:
        return []
    base_len = len(primero)
    numeros = [int(primero)]
    for p in partes[1:]:
        p_digits = re.sub(r'\D', '', p)
        if not p_digits:
            continue
        if len(p_digits) < base_len:
            prefijo = primero[:base_len - len(p_digits)]
            numeros.append(int(prefijo + p_digits))
        else:
            numeros.append(int(p_digits))
    return ['C' + str(n) for n in numeros]


def limpiar_num(s):
    if not s: return None
    s = str(s).strip().replace(' ', '')
    if not s: return None
    if re.match(r'^\d+\.\d{3}$', s):
        entero, dec = s.split('.')
        dec_limpio = dec.rstrip('0')
        s = entero + '.' + dec_limpio if dec_limpio else entero
    elif ',' in s and re.match(r'^\d{1,3}(\.\d{3})+(,\d+)?$', s):
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace(',', '.')
    try: return '{:.2f}'.format(float(s))
    except Exception: return s


def formatear_miles(valor):
    """Convierte un numero (string con punto decimal, ej '21692.00') al formato
    con separador de miles por punto y decimales con coma (ej '21.692,00'),
    usado en el certificado de Mexico."""
    try:
        f = float(valor)
    except (TypeError, ValueError):
        return valor
    entero, dec = '{:.2f}'.format(f).split('.')
    signo = ''
    if entero.startswith('-'):
        signo = '-'
        entero = entero[1:]
    entero_fmt = '{:,}'.format(int(entero)).replace(',', '.')
    return signo + entero_fmt + ',' + dec


# ── REMITO (fitz) ────────────────────────────────────────────────────────────

def leer_remito(pdf_bytes):
    doc = fitz.open(stream=pdf_bytes, filetype='pdf')
    texto = ''
    for page in doc: texto += page.get_text()
    doc.close()
    datos = {}
    m_vuelo = re.search(r'Buque/Aerol[íi]nea[:\s]+([^\n\r]+)', texto, re.IGNORECASE)
    if m_vuelo:
        transport = m_vuelo.group(1).strip()
        if '-' in transport:
            partes = transport.split('-')
            if len(partes) == 2 and re.match(r'[A-Z]{2}\d+', partes[1]):
                datos['transporte'] = partes[1]
                datos['tipo_transporte'] = 'aereo'
            else:
                datos['transporte'] = transport
                datos['tipo_transporte'] = 'maritimo'
        else:
            datos['transporte'] = transport
            datos['tipo_transporte'] = 'maritimo'
    m_cont = re.search(r'CONTAINER[:\s]*([A-Z]{4}\d{6,7}-?\d?)', texto, re.IGNORECASE)
    datos['contenedor'] = m_cont.group(1).strip() if m_cont else None
    m_ps = re.search(r'P\.S\.[:\s]+([A-Z0-9/]+)', texto)
    m_pa = re.search(r'P\.A\.[:\s]+([A-Z]{2,3}\s?\d{4,8})', texto)
    datos['precinto_senasa'] = m_ps.group(1).strip() if m_ps else None
    datos['precinto_afip']   = m_pa.group(1).strip() if m_pa else None
    m_contra = re.search(r'CONTRAMARCA[:\s]+([^\n\r]+)', texto, re.IGNORECASE)
    contra = m_contra.group(1).strip() if m_contra else ''
    datos['contramarca'] = contra if contra else None
    m_pallets = re.search(r'EN\s+(\d+)\s+PALLETS?', texto, re.IGNORECASE)
    datos['pallets'] = m_pallets.group(1) if m_pallets else None
    m_tot_cajas = re.search(r'Total General\s+(\d[\d\.]*)', texto)
    m_tot_neto  = re.search(r'PESO NETO TOTAL[:\s]+([\d\.,]+)', texto)
    m_tot_bruto = re.search(r'PESO BRUTO TOTAL[:\s]+([\d\.,]+)', texto)
    datos['total_cajas'] = m_tot_cajas.group(1).replace('.', '') if m_tot_cajas else None
    datos['total_neto']  = limpiar_num(m_tot_neto.group(1)) if m_tot_neto else None
    datos['total_bruto'] = limpiar_num(m_tot_bruto.group(1)) if m_tot_bruto else None
    productos = []
    lineas = texto.split('\n')
    i = 0
    while i < len(lineas):
        linea = lineas[i].strip()
        if re.match(r'^CD\d+$', linea):
            codigo    = linea
            desc      = lineas[i+1].strip() if i+1 < len(lineas) else ''
            cajas     = lineas[i+2].strip() if i+2 < len(lineas) else ''
            neto_raw  = lineas[i+4].strip() if i+4 < len(lineas) else ''
            bruto_raw = lineas[i+5].strip() if i+5 < len(lineas) else ''
            nombre_es = buscar_nombre_es_remito(desc)
            productos.append({
                'codigo': codigo, 'nombre_es': nombre_es, 'nombre_en': '',
                'desc_original': desc,
                'cajas': cajas, 'neto': limpiar_num(neto_raw), 'bruto': limpiar_num(bruto_raw),
            })
            i += 6
        else:
            i += 1
    datos['productos'] = productos

    # Detectar congelado desde observaciones del remito
    if re.search(r'CONGELAD', texto, re.IGNORECASE):
        datos['es_congelado'] = True
    else:
        datos['es_congelado'] = False

    return datos


# ── SANITARIO PROVISORIO (OCR) ───────────────────────────────────────────────

def ocr_pdf(pdf_bytes):
    texto = ''
    with tempfile.TemporaryDirectory() as tmpdir:
        pdf_path = os.path.join(tmpdir, 'input.pdf')
        with open(pdf_path, 'wb') as f: f.write(pdf_bytes)
        out_prefix = os.path.join(tmpdir, 'page')
        subprocess.run(['pdftoppm', '-r', '300', '-l', '1', '-jpeg', pdf_path, out_prefix],
            check=True, capture_output=True)
        archivos = sorted([f for f in os.listdir(tmpdir) if f.startswith('page') and f.endswith('.jpg')])
        for nombre in archivos:
            img = Image.open(os.path.join(tmpdir, nombre))
            texto += pytesseract.image_to_string(img, lang='spa') + '\n'
            img.close()
    return texto


def leer_sanitario_provisorio(pdf_bytes):
    texto = ocr_pdf(pdf_bytes)
    datos = {}
    m_kg = re.search(r'EN\s+(\d+)\s+PALLETS?[:\s]+([\d\.,]+)', texto, re.IGNORECASE)
    if m_kg:
        datos['pallets_prov'] = m_kg.group(1)
        datos['kg_pallets']   = limpiar_num(m_kg.group(2))
    else:
        datos['kg_pallets'] = None
    # Detectar congelado vs enfriado
    if re.search(r'CONGELAD', texto, re.IGNORECASE):
        datos['es_congelado'] = True
    else:
        datos['es_congelado'] = False
    datos['fecha_emision'] = datetime.datetime.now().strftime('%d/%m/%Y')

    # Contramarca por linea, en el orden en que aparecen (formato USA con anexo,
    # ej. "55 ... - C208 ( Fecha de Faena: ... )"). El OCR a veces confunde la
    # 'C' del codigo con otro caracter (0, 9, etc.) - se ignora ese caracter y
    # se reconstruye siempre con 'C' + los digitos. El numero de cajas al
    # principio de la linea es opcional porque el OCR a veces lo pierde del
    # todo (ej. una linea entera sin numero visible). La fecha de faena de esta
    # linea NO se usa (viene poco confiable del OCR) - se usa la del piqueo por
    # producto en su lugar; esto solo ancla el match a filas de producto reales
    # (evita matchear el resumen "CONTRAMARCA:C208/C209/...").
    lineas_usa = []
    patron_linea = re.compile(
        r'^\s*(\d+)?([^\n]*?)-\s*[A-Za-z0-9](\d{2,4})\s*\(\s*Fecha de Faena',
        re.IGNORECASE | re.MULTILINE
    )
    for m in patron_linea.finditer(texto):
        lineas_usa.append({'cajas': m.group(1), 'texto': (m.group(2) or '').upper(), 'contramarca': 'C' + m.group(3)})
    datos['lineas_usa'] = lineas_usa

    return datos


# ── MAPAS NOMBRE ─────────────────────────────────────────────────────────────

MAPA_EN = {
    'BOLA DE LOMO':              'KNUCKLE',
    'CUADRADA':                  'OUTSIDE FLAT',
    'LOMO SIN CORDON':           'TENDERLOIN CHAIN OFF',
    'LOMO S/ CORDON':            'TENDERLOIN CHAIN OFF',
    'LOMO SC':                   'TENDERLOIN CHAIN OFF',
    'NALGA DE ADENTRO CON TAPA': 'TOPSIDE CAP ON',
    'NALGA CON TAPA':            'TOPSIDE CAP ON',
    'NALGA SIN TAPA':            'TOPSIDE CAP OFF',
    'NALGA':                     'TOPSIDE',
    'CARNAZA DE PALETA':         'BOLAR BLADE',
    'BIFE ANGOSTO':              'STRIPLOIN',
    'BIFE ANCHO SIN TAPA':       'RIBEYE',
    'BIFE ANCHO ST':             'RIBEYE',
    'BIFE ANCHO':                'RIB EYE',
    'COLITA DE CUADRIL':         'TRI-TIP',
    'CORAZON DE CUADRIL':        'HEART OF RUMP',
    'MARUCHA':                   'OYSTER BLADE',
    'ASADO SIN HUESO':           'SHORT RIB MEAT',
    'PECHO':                     'BRISKET POINT END',
    'TAPA DE BIFE ANCHO':        'RIB CAP',
    'TAPA DE CUADRIL':           'RUMP CAP',
    'BIFE DE VACIO GRANDE':      'FLAP MEAT',
    'BIFE DE VACIO':             'FLANK',
    'PECETO':                    'EYE ROUND',
    'AGUJA':                     'CHUCK',
    'CHINGOLO':                  'CHUCK TENDER',
    'CORAZON DE PALETA':         'SHOULDER CLOD HEART',
    'BRAZUELO':                  'CONICAL MUSCLE',
    'COGOTE':                    'NECK',
    'PECHO PEDO':                'BRISKET POINT END',
    'CABEZA DE LOMO':            'TENDERLOIN BUTT',
    'BIFE ANGOSTO CON HUESO':    'SHORTLOIN',
    'CARNAZA':                   'BOLAR BLADE',
}

NOMBRES_PRODUCTO = sorted(MAPA_EN.keys(), key=len, reverse=True)


def buscar_nombre_es_remito(desc):
    d = desc.upper()
    for nombre in NOMBRES_PRODUCTO:
        if nombre in d:
            lbs_m = re.search(r'(\d/\d\s*LBS|\+\s*\d\s*LBS|\+5\s*LBS)', desc, re.IGNORECASE)
            if lbs_m and 'LOMO' in nombre:
                return nombre + ' ' + lbs_m.group(1).strip()
            return nombre
    return desc


def buscar_nombre_en(nombre_es):
    n = nombre_es.upper()
    for clave, en in sorted(MAPA_EN.items(), key=lambda x: len(x[0]), reverse=True):
        if clave in n: return en
    return ''


def armar_nombre_bilingue(nombre_es, nombre_en):
    es = nombre_es.strip().upper()
    en = (nombre_en or '').strip().upper()
    if en and en != es: return es + '/ ' + en
    return es


# ── NOMBRES ESPECIFICOS MÉXICO (corte / "pulpa" / ingles) ────────────────────
# Tabla fija por tipo de corte, requerida por la certificacion mexicana.
# No todos los cortes llevan calificador de "pulpa" (ver TAPA DE CUADRIL,
# BIFE ANGOSTO, BIFE ANCHO en los ejemplos - no llevan).
MAPA_MEXICO = {
    'NALGA DE AFUERA':  {'es': 'NALGA DE AFUERA CT', 'pulpa': 'PULPA BLANCA', 'en': 'BEEF GOOSENECK'},
    'NALGA CON TAPA':   {'es': 'NALGA CON TAPA',      'pulpa': 'PULPA NEGRA', 'en': 'BEEF TOP (INSIDE) ROUND'},
    'BOLA DE LOMO':     {'es': 'BOLA DE LOMO',        'pulpa': 'PULPA BOLA',  'en': 'BONELESS BEEF KNUCKLE'},
    'TAPA DE CUADRIL':  {'es': 'TAPA DE CUADRIL',     'pulpa': None,          'en': 'BONELESS BEEF RUMP CAP'},
    'BIFE ANGOSTO':     {'es': 'BIFE ANGOSTO CC',     'pulpa': None,          'en': 'BONELESS BEEF NEW YORK'},
    'BIFE ANCHO':       {'es': 'BIFE ANCHO',          'pulpa': None,          'en': 'BONELESS BEEF RIB EYE'},
}
CLAVES_MEXICO = sorted(MAPA_MEXICO.keys(), key=len, reverse=True)


def buscar_info_mexico(desc_original):
    """Busca el corte dentro de la descripcion cruda del remito (ej.
    'NALGA DE AFUERA C/TORTGUITA (MEX) GF') y devuelve su info de Mexico,
    o None si no esta en la tabla (corte nuevo, no mapeado todavia)."""
    d = (desc_original or '').upper()
    for clave in CLAVES_MEXICO:
        if clave in d:
            return MAPA_MEXICO[clave]
    return None


def armar_nombre_mexico(prod):
    """Arma el nombre de 3 partes 'ES/ PULPA / EN' (o 2 partes 'ES / EN' si
    el corte no lleva pulpa) para el certificado de Mexico. Si el corte no
    esta en MAPA_MEXICO, cae al nombre bilingue generico (mejor avisar con
    una alerta manualmente que dejar la celda vacia)."""
    info = buscar_info_mexico(prod.get('desc_original', ''))
    if info is None:
        return armar_nombre_bilingue(prod.get('nombre_es', ''), prod.get('nombre_en', ''))
    es, pulpa, en = info['es'], info['pulpa'], info['en']
    if pulpa:
        return es + '/ ' + pulpa + ' / ' + en
    return es + ' / ' + en


# ── XML HELPERS ──────────────────────────────────────────────────────────────

def get_trs(xml):
    return list(re.finditer(r'<w:tr[ >]', xml))


def get_fila_xml(xml, trs, idx):
    ini = trs[idx].start()
    fin = trs[idx + 1].start() if idx + 1 < len(trs) else len(xml)
    return xml[ini:fin], ini, fin


def _reemplazar_celda(xml_fila, celda_idx, nuevo_texto):
    celda_starts = [m.start() for m in re.finditer(r'<w:tc>', xml_fila)]
    celda_ends   = [m.start() for m in re.finditer(r'</w:tc>', xml_fila)]
    if celda_idx >= len(celda_starts): return xml_fila
    bloque = xml_fila[celda_starts[celda_idx]:celda_ends[celda_idx]]
    textos = re.findall(r'<w:t[^>]*>[^<]*</w:t>', bloque)
    if not textos: return xml_fila
    primer   = textos[0]
    tag_open = re.match(r'<w:t[^>]*>', primer).group()
    nuevo_bloque = bloque.replace(primer, tag_open + nuevo_texto + '</w:t>', 1)
    for t in textos[1:]:
        tag2 = re.match(r'<w:t[^>]*>', t).group()
        nuevo_bloque = nuevo_bloque.replace(t, tag2 + '</w:t>', 1)
    return xml_fila[:celda_starts[celda_idx]] + nuevo_bloque + xml_fila[celda_ends[celda_idx]:]


def _construir_fila(fila_modelo, cajas, nombre_bi, neto, bruto, neto_celda, bruto_celda,
                     lotes=None, lotes_celda=None):
    nueva = fila_modelo
    nueva = _reemplazar_celda(nueva, 0, str(cajas))
    nueva = _reemplazar_celda(nueva, 1, nombre_bi)
    if lotes_celda is not None:
        nueva = _reemplazar_celda(nueva, lotes_celda, str(lotes or ''))
    nueva = _reemplazar_celda(nueva, neto_celda, str(neto))
    nueva = _reemplazar_celda(nueva, bruto_celda, str(bruto))
    return nueva


def _get_fila_por_contenido(xml, trs, texto_clave):
    for i, m in enumerate(trs):
        ini = m.start()
        fin = trs[i+1].start() if i+1 < len(trs) else len(xml)
        if texto_clave in xml[ini:fin]:
            return xml[ini:fin], ini, fin, i
    return None, None, None, None


def _reemplazar_pallets_en_fila(fila_xml, pallets, kg_pallets):
    fila_xml = re.sub(r'(ACONDICIONADO EN\s*)\d+(\s*PALLET)', r'\g<1>' + str(pallets) + r'\2', fila_xml)
    fila_xml = re.sub(r'(ACONDITIONED IN\s*)\d+(\s*PALLET)', r'\g<1>' + str(pallets) + r'\2', fila_xml)
    fila_xml = fila_xml.replace('<w:t>' + str(pallets) + '</w:t>', '<w:t>' + str(pallets) + '</w:t>', 1)
    # Reemplazar numero separado - buscar w:t con solo digitos pequeños
    for viejo in ['<w:t>1</w:t>', '<w:t>4</w:t>', '<w:t>14</w:t>', '<w:t>19</w:t>', '<w:t>20</w:t>']:
        if viejo in fila_xml:
            fila_xml = fila_xml.replace(viejo, '<w:t>' + str(pallets) + '</w:t>', 1)
            break
    if kg_pallets:
        # Caso generico: el numero de KGS esta en el mismo run que el texto "KGS)" (ej. Mexico)
        nueva_fila, n = re.subn(r'[\d\.]+(\s*KGS\))', str(kg_pallets) + r'\1', fila_xml, count=1)
        if n:
            fila_xml = nueva_fila
        else:
            # Fallback: numero de KGS aislado en su propio w:t (plantillas viejas)
            for viejo_kg in ['<w:t>32.44</w:t>', '<w:t>151.77</w:t>', '<w:t>664.35</w:t>']:
                if viejo_kg in fila_xml:
                    fila_xml = fila_xml.replace(viejo_kg, '<w:t>' + str(kg_pallets) + '</w:t>')
                    break
    return fila_xml


def _reemplazar_bloque_productos(xml, trs, primera_idx, total_idx_fallback,
                                  productos, total_cajas, total_neto, total_bruto,
                                  pallets, kg_pallets, neto_celda=6, bruto_celda=7,
                                  lotes_celda=None, sumar_pallet_a_bruto=True,
                                  total_replacer=None, armar_nombre_func=None):
    fila_pal, ini_pal, fin_pal, idx_pal = _get_fila_por_contenido(xml, trs, 'ACONDICIONADO EN')
    total_idx = (idx_pal + 1) if idx_pal is not None else total_idx_fallback

    fila_modelo, ini_mod, _ = get_fila_xml(xml, trs, primera_idx)
    fila_total, ini_tot, fin_tot = get_fila_xml(xml, trs, total_idx)

    if armar_nombre_func is None:
        armar_nombre_func = lambda prod: armar_nombre_bilingue(prod.get('nombre_es', ''), prod.get('nombre_en', ''))

    nuevas_filas = ''
    for prod in productos:
        nombre_bi = armar_nombre_func(prod)
        nuevas_filas += _construir_fila(
            fila_modelo, prod.get('cajas', ''), nombre_bi,
            prod.get('neto', ''), prod.get('bruto', ''), neto_celda, bruto_celda,
            lotes=prod.get('lotes', ''), lotes_celda=lotes_celda
        )

    nueva_pal = _reemplazar_pallets_en_fila(fila_pal, pallets, kg_pallets) if fila_pal else ''

    if sumar_pallet_a_bruto:
        try:
            total_bruto_final = '{:.2f}'.format(float(total_bruto) + float(kg_pallets or 0))
        except Exception:
            total_bruto_final = total_bruto
    else:
        total_bruto_final = total_bruto

    if total_replacer is not None:
        nueva_total = total_replacer(fila_total, total_cajas, total_neto, total_bruto_final)
    else:
        nums_tot = re.findall(r'<w:t[^>]*>(\d[\d\.]*)</w:t>', fila_total)
        nueva_total = fila_total
        if len(nums_tot) >= 3:
            nueva_total = nueva_total.replace('>' + nums_tot[0] + '<', '>' + str(total_cajas) + '<', 1)
            nueva_total = nueva_total.replace('>' + nums_tot[1] + '<', '>' + str(total_neto) + '<', 1)
            nueva_total = nueva_total.replace('>' + nums_tot[2] + '<', '>' + str(total_bruto_final) + '<', 1)

    xml_nuevo = xml[:ini_mod] + nuevas_filas + nueva_pal + nueva_total + xml[fin_tot:]
    return xml_nuevo


def fmt_fecha_al_to(f):
    if f and ' al ' in f.lower():
        partes = re.split(r'\s+al\s+', f, flags=re.IGNORECASE)
        return partes[0] + ' AL/TO ' + partes[1]
    return f or ''


def fmt_fecha_al(f):
    if f and ' al ' in f.lower():
        partes = re.split(r'\s+al\s+', f, flags=re.IGNORECASE)
        return partes[0] + ' AL ' + partes[1]
    return f or ''


def fmt_fecha_al_to_usa(f):
    """'dd/mm/yyyy al dd/mm/yyyy' -> 'dd/mm/yyyy al/to dd/mm/yyyy' (formato USA, minuscula)."""
    if f and ' al ' in f.lower():
        partes = re.split(r'\s+al\s+', f, flags=re.IGNORECASE)
        return partes[0].strip() + ' al/to ' + partes[1].strip()
    return f or ''


def fecha_a_lote_usa(f):
    """'dd/mm/yyyy al dd/mm/yyyy' -> 'YYYYMMDD al/to YYYYMMDD' (mismo rango que fecha de produccion, Lote de USA)."""
    if not f or ' al ' not in f.lower():
        return f or ''
    partes = re.split(r'\s+al\s+', f, flags=re.IGNORECASE)
    salida = []
    for p in partes:
        m = re.match(r'(\d{2})/(\d{2})/(\d{4})', p.strip())
        salida.append(m.group(3) + m.group(2) + m.group(1) if m else p.strip())
    return ' al/to '.join(salida)


def kg_a_lbs(kg):
    try:
        return '{:.2f}'.format(float(str(kg).replace(',', '.')) * 2.20462)
    except (TypeError, ValueError):
        return ''


def _merge_runs_xml(xml):
    """Fusiona <w:r> adyacentes con el mismo <w:rPr> dentro de cada parrafo,
    concatenando sus <w:t>. Word fragmenta el texto en runs distintos (marcas
    de revision, corrector ortografico), lo que rompe los reemplazos simples
    basados en substring (ej. 'VAPOR/VESSEL:  NOMBRE' guardado en 3 runs
    separados). Solo fusiona runs de texto simple (un unico <w:t>, sin tabs,
    saltos de linea u otros elementos) para no arriesgar contenido complejo."""
    xml = re.sub(r'<w:proofErr[^/]*/>', '', xml)  # el corrector ortografico bloquea la fusion de runs adyacentes

    def _procesar_parrafo(m):
        parrafo = m.group(0)
        cambiado = True
        while cambiado:
            cambiado = False
            runs = list(re.finditer(r'<w:r(?:\s[^>]*)?>(.*?)</w:r>', parrafo, re.DOTALL))
            for i in range(len(runs) - 1):
                r1, r2 = runs[i], runs[i + 1]
                if parrafo[r1.end():r2.start()]:
                    continue  # no son estrictamente adyacentes
                rpr1 = re.search(r'<w:rPr>.*?</w:rPr>', r1.group(1), re.DOTALL)
                rpr2 = re.search(r'<w:rPr>.*?</w:rPr>', r2.group(1), re.DOTALL)
                if (rpr1.group(0) if rpr1 else '') != (rpr2.group(0) if rpr2 else ''):
                    continue
                t1 = re.findall(r'<w:t[^>]*>([^<]*)</w:t>', r1.group(1))
                t2 = re.findall(r'<w:t[^>]*>([^<]*)</w:t>', r2.group(1))
                # Solo fusionar runs de un unico <w:t> simple (evita tabs/breaks/drawings)
                if len(t1) != 1 or len(t2) != 1:
                    continue
                if r1.group(1).count('<w:t') != 1 or r2.group(1).count('<w:t') != 1:
                    continue
                texto_unido = t1[0] + t2[0]
                rpr_txt = rpr1.group(0) if rpr1 else ''
                nuevo_run = '<w:r>' + rpr_txt + '<w:t xml:space="preserve">' + texto_unido + '</w:t></w:r>'
                parrafo = parrafo[:r1.start()] + nuevo_run + parrafo[r2.end():]
                cambiado = True
                break
        return parrafo

    return re.sub(r'<w:p(?:\s[^>]*)?>.*?</w:p>', _procesar_parrafo, xml, flags=re.DOTALL)


def generar_sanitario(docx_bytes, datos, tipo_via, destino):
    alertas = []
    with zipfile.ZipFile(io.BytesIO(docx_bytes), 'r') as z:
        archivos = {n: z.read(n) for n in z.namelist()}
    xml = archivos['word/document.xml'].decode('utf-8')
    xml = _merge_runs_xml(xml)
    if destino == 'singapur':
        if tipo_via == 'aereo':
            xml, al = _gen_singapur_aereo(xml, datos)
        else:
            xml, al = _gen_singapur_maritimo(xml, datos)
    elif destino == 'mexico':
        xml, al = _gen_mexico_maritimo(xml, datos)
    elif destino == 'usawclass':
        xml, al = _gen_usa_wclass(xml, datos)
    elif destino == 'usaorleans':
        xml, al = _gen_usa_orleans(xml, datos)
    else:
        if tipo_via == 'aereo':
            xml, al = _gen_malasia_aereo(xml, datos)
        else:
            xml, al = _gen_malasia_maritimo(xml, datos)
    alertas.extend(al)

    # Red de seguridad: si por algun bug el XML quedo mal formado, no entregar
    # un .docx roto (Word no lo puede ni abrir) - mejor fallar con un error claro.
    try:
        from xml.etree import ElementTree as ET
        ET.fromstring(xml)
    except ET.ParseError as e:
        raise ValueError(
            'El documento generado quedo con XML invalido (' + str(e) + '). '
            'No se genero el archivo para evitar entregar un .docx corrupto - '
            'avisar para revisar el generador de este destino.'
        )

    archivos['word/document.xml'] = xml.encode('utf-8')
    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as z:
        for n, d in archivos.items(): z.writestr(n, d)
    out.seek(0)
    return out.read(), alertas



def _reemplazar_fechas(xml, trs, f_faena, f_prod, f_venc, fmt_func):
    """Busca filas con I.13/I.14/I.15 y reemplaza solo la parte de fecha,
    preservando los titulos (I.13.Fecha de faena/, Date of slaughter, etc.)"""

    def _get_celdas(fila):
        starts = [m.start() for m in re.finditer(r'<w:tc>', fila)]
        ends   = [m.start() for m in re.finditer(r'</w:tc>', fila)]
        return [(s, e) for s, e in zip(starts, ends)]

    def _reemplazar_fecha_en_celda(fila, cs, ce, nuevo):
        bloque = fila[cs:ce]
        # Encontrar todos los w:t con su posicion
        wts = list(re.finditer(r'<w:t[^>]*>[^<]*</w:t>', bloque))
        if not wts: return fila

        # Encontrar el primer w:t que contiene digitos de fecha (dd o /mm o /yyyy)
        # Los titulos son texto como "I.13.Fecha de faena/" y "Date of slaughter"
        # La fecha empieza cuando aparece un fragmento con solo digitos o slash
        # Reconstruir texto completo de la celda para encontrar donde empieza la fecha
        txt_completo = ''.join(re.search(r'<w:t[^>]*>([^<]*)</w:t>', wt.group()).group(1) for wt in wts)
        # Buscar posicion del primer dd/ en el texto completo
        m_fecha = re.search(r'\d{2}/', txt_completo)
        if not m_fecha:
            return fila

        # Encontrar que w:t corresponde a esa posicion
        fecha_inicio_idx = None
        pos_acum = 0
        for idx, wt in enumerate(wts):
            txt_wt = re.search(r'<w:t[^>]*>([^<]*)</w:t>', wt.group()).group(1)
            if pos_acum + len(txt_wt) > m_fecha.start():
                fecha_inicio_idx = idx
                break
            pos_acum += len(txt_wt)

        if fecha_inicio_idx is None: return fila

        # Reemplazar desde fecha_inicio_idx en adelante
        nuevo_bloque = bloque
        offset = 0
        for idx, wt in enumerate(wts):
            if idx < fecha_inicio_idx: continue
            tag = re.match(r'<w:t[^>]*>', wt.group()).group()
            old_wt = wt.group()
            pos = nuevo_bloque.find(old_wt, offset)
            if idx == fecha_inicio_idx:
                # Primer fragmento de fecha: poner el valor nuevo con preserve
                new_wt = '<w:t xml:space="preserve">' + nuevo + '</w:t>'
            else:
                # Fragmentos siguientes: vaciar
                new_wt = tag + '</w:t>'
            nuevo_bloque = nuevo_bloque[:pos] + new_wt + nuevo_bloque[pos + len(old_wt):]
            offset = pos + len(new_wt)

        return fila[:cs] + nuevo_bloque + fila[ce:]

    fechas_a_reemplazar = [(f_faena, fmt_func), (f_prod, fmt_func), (f_venc, fmt_func)]
    fecha_idx = 0

    for i, m in enumerate(trs):
        if fecha_idx >= 3: break
        ini = m.start()
        fin = trs[i+1].start() if i+1 < len(trs) else len(xml)
        fila = xml[ini:fin]

        if not ('I.13' in fila or 'I.14' in fila or 'I.15' in fila):
            continue

        celdas = _get_celdas(fila)
        nueva_fila = fila
        offset = 0

        for cs, ce in celdas:
            if fecha_idx >= 3: break
            bloque = nueva_fila[cs+offset:ce+offset]
            txt_celda = ''.join(re.findall(r'<w:t[^>]*>([^<]*)</w:t>', bloque))
            if re.search(r'\d{2}/\d{2}/\d{4}', txt_celda):
                f_nueva, fmt = fechas_a_reemplazar[fecha_idx]
                if f_nueva:
                    fila_antes = nueva_fila
                    nueva_fila = _reemplazar_fecha_en_celda(nueva_fila, cs+offset, ce+offset, fmt(f_nueva))
                    offset += len(nueva_fila) - len(fila_antes)
                fecha_idx += 1

        xml = xml[:ini] + nueva_fila + xml[fin:]

    return xml

# ── MALASIA AÉREO ────────────────────────────────────────────────────────────

def _gen_malasia_aereo(xml, datos):
    alertas = []
    trs = get_trs(xml)
    xml = _reemplazar_bloque_productos(
        xml, trs, primera_idx=5, total_idx_fallback=19,
        productos=datos.get('productos', []),
        total_cajas=datos.get('total_cajas',''), total_neto=datos.get('total_neto',''), total_bruto=datos.get('total_bruto',''),
        pallets=datos.get('pallets','1'), kg_pallets=datos.get('kg_pallets',''),
        neto_celda=6, bruto_celda=7
    )
    f_faena = datos.get('fecha_faena','')
    f_prod  = datos.get('fecha_produccion','')
    f_venc  = datos.get('fecha_vencimiento','')
    trs2 = get_trs(xml)
    xml = _reemplazar_fechas(xml, trs2, f_faena, f_prod, f_venc, fmt_fecha_al_to)
    transporte = datos.get('transporte','')
    if transporte: xml = xml.replace('>VUELO / FLIGHT: EK248<', '>VUELO / FLIGHT: ' + transporte + '<')
    xml = _set_temperatura_singapur(xml, datos.get('es_congelado', False), tipo_via='aereo')
    fecha_emi = datos.get('fecha_emision') or datetime.datetime.now().strftime('%d/%m/%Y')
    try: dia, mes, anio = fecha_emi.split('/')
    except: dia = mes = anio = ''; alertas.append('Fecha emision no parseada')
    xml = xml.replace('>2026<',  '>' + anio + '<', 1)
    xml = xml.replace('>01 <',   '>' + mes + ' <', 1)
    xml = xml.replace('>23<',    '>' + dia + '<', 1)
    return xml, alertas


# ── MALASIA MARÍTIMO ─────────────────────────────────────────────────────────

def _gen_malasia_maritimo(xml, datos):
    alertas = []
    trs = get_trs(xml)
    xml = _reemplazar_bloque_productos(
        xml, trs, primera_idx=5, total_idx_fallback=15,
        productos=datos.get('productos', []),
        total_cajas=datos.get('total_cajas',''), total_neto=datos.get('total_neto',''), total_bruto=datos.get('total_bruto',''),
        pallets=datos.get('pallets','1'), kg_pallets=datos.get('kg_pallets',''),
        neto_celda=6, bruto_celda=7
    )
    f_faena = datos.get('fecha_faena','')
    f_prod  = datos.get('fecha_produccion','')
    f_venc  = datos.get('fecha_vencimiento','')
    trs2 = get_trs(xml)
    xml = _reemplazar_fechas(xml, trs2, f_faena, f_prod, f_venc, fmt_fecha_al)
    transporte = datos.get('transporte','')
    if transporte: xml = xml.replace('>VAPOR / VESSEL: TIGER PLATA<', '>VAPOR / VESSEL: ' + transporte + '<')
    contenedor    = datos.get('contenedor','')
    precinto_afip = datos.get('precinto_afip','')
    if contenedor:    xml = xml.replace('>TCLU129408-4<', '>' + contenedor + '<')
    if precinto_afip: xml = xml.replace('>BAH79585<',    '>' + precinto_afip + '<')
    if not contenedor:    alertas.append('Contenedor no encontrado - completar manualmente')
    if not precinto_afip: alertas.append('Precinto AFIP no encontrado - completar manualmente')
    xml = _set_temperatura_singapur(xml, datos.get('es_congelado', False), tipo_via='aereo')
    fecha_emi = datos.get('fecha_emision') or datetime.datetime.now().strftime('%d/%m/%Y')
    try: dia, mes, anio = fecha_emi.split('/')
    except: dia = mes = anio = ''; alertas.append('Fecha emision no parseada')
    xml = xml.replace('>:      2026<', '>:      ' + anio + '<')
    xml = re.sub(r'>\)\s+\d{2}\s+<', '>)         ' + mes + ' <', xml, count=1)
    xml = xml.replace('>14<', '>' + dia + '<', 1)
    return xml, alertas


# ── SINGAPUR AÉREO ───────────────────────────────────────────────────────────

def _gen_singapur_aereo(xml, datos):
    alertas = []
    trs = get_trs(xml)
    xml = _reemplazar_bloque_productos(
        xml, trs, primera_idx=5, total_idx_fallback=17,
        productos=datos.get('productos', []),
        total_cajas=datos.get('total_cajas',''), total_neto=datos.get('total_neto',''), total_bruto=datos.get('total_bruto',''),
        pallets=datos.get('pallets','1'), kg_pallets=datos.get('kg_pallets',''),
        neto_celda=5, bruto_celda=6
    )
    # Fechas
    f_faena = datos.get('fecha_faena','')
    f_prod  = datos.get('fecha_produccion','')
    f_venc  = datos.get('fecha_vencimiento','')
    trs2 = get_trs(xml)
    xml = _reemplazar_fechas(xml, trs2, f_faena, f_prod, f_venc, fmt_fecha_al_to)
    # Transporte (vuelo)
    transporte = datos.get('transporte','')
    if transporte: xml = xml.replace('>: LX093<', '>: ' + transporte + '<')
    # Temperatura
    es_congelado = datos.get('es_congelado', False)
    xml = _set_temperatura_singapur(xml, es_congelado, tipo_via='aereo')
    # Consignatario
    xml = xml.replace('>FOODIE MARKET PLACE PTE. LTD<', '>FOODIE MARKET PLACE PTE. LTD<')  # placeholder
    # Fecha emision
    fecha_emi = datos.get('fecha_emision') or datetime.datetime.now().strftime('%d/%m/%Y')
    xml = xml.replace('>04/06/2026<', '>' + fecha_emi + '<')
    return xml, alertas


# ── SINGAPUR MARÍTIMO ────────────────────────────────────────────────────────

def _gen_singapur_maritimo(xml, datos):
    alertas = []
    trs = get_trs(xml)
    xml = _reemplazar_bloque_productos(
        xml, trs, primera_idx=5, total_idx_fallback=10,
        productos=datos.get('productos', []),
        total_cajas=datos.get('total_cajas',''), total_neto=datos.get('total_neto',''), total_bruto=datos.get('total_bruto',''),
        pallets=datos.get('pallets','1'), kg_pallets=datos.get('kg_pallets',''),
        neto_celda=6, bruto_celda=7
    )
    # Fechas
    f_faena = datos.get('fecha_faena','')
    f_prod  = datos.get('fecha_produccion','')
    f_venc  = datos.get('fecha_vencimiento','')
    trs2 = get_trs(xml)
    xml = _reemplazar_fechas(xml, trs2, f_faena, f_prod, f_venc, fmt_fecha_al_to)
    # Transporte (barco)
    transporte = datos.get('transporte','')
    if transporte: xml = xml.replace('>: SAN ANTONIO MAERSK<', '>: ' + transporte + '<')
    # Contenedor y precinto
    contenedor    = datos.get('contenedor','')
    precinto_afip = datos.get('precinto_afip','')
    if contenedor:    xml = xml.replace('>MNBU9179760<', '>' + contenedor + '<')
    if precinto_afip: xml = xml.replace('>BAH66389<',   '>' + precinto_afip + '<')
    if not contenedor:    alertas.append('Contenedor no encontrado - completar manualmente')
    if not precinto_afip: alertas.append('Precinto AFIP no encontrado - completar manualmente')
    # Temperatura
    es_congelado = datos.get('es_congelado', False)
    xml = _set_temperatura_singapur(xml, es_congelado, tipo_via='maritimo')
    # Fecha emision
    fecha_emi = datos.get('fecha_emision') or datetime.datetime.now().strftime('%d/%m/%Y')
    xml = xml.replace('>14/04/2026<', '>' + fecha_emi + '<')
    return xml, alertas


# ── MÉXICO MARÍTIMO ──────────────────────────────────────────────────────────

def _reemplazar_total_celdas(fila_total, total_cajas, total_neto, total_bruto,
                              cajas_celda=0, neto_celda=2, bruto_celda=3):
    """Reemplaza la fila de totales por indice de celda en vez de buscar numeros
    sueltos en el XML. Necesario cuando los totales usan formato con coma
    decimal (ej. '21.692,00'), que no matchea como numero simple."""
    nueva = fila_total
    nueva = _reemplazar_celda(nueva, cajas_celda, str(total_cajas))
    nueva = _reemplazar_celda(nueva, neto_celda, str(total_neto))
    nueva = _reemplazar_celda(nueva, bruto_celda, str(total_bruto))
    return nueva


def _gen_mexico_maritimo(xml, datos):
    alertas = []
    trs = get_trs(xml)

    total_neto_fmt  = formatear_miles(datos.get('total_neto', ''))
    total_bruto_fmt = formatear_miles(datos.get('total_bruto', ''))

    for prod in datos.get('productos', []):
        if buscar_info_mexico(prod.get('desc_original', '')) is None:
            alertas.append(
                'Corte "' + prod.get('desc_original', prod.get('nombre_es', ''))
                + '" no esta en la tabla de nombres de Mexico - se uso el nombre generico, revisar manualmente'
            )

    xml = _reemplazar_bloque_productos(
        xml, trs, primera_idx=5, total_idx_fallback=13,
        productos=datos.get('productos', []),
        total_cajas=datos.get('total_cajas', ''), total_neto=total_neto_fmt, total_bruto=total_bruto_fmt,
        pallets=datos.get('pallets', '1'), kg_pallets=datos.get('kg_pallets', ''),
        neto_celda=6, bruto_celda=7,
        lotes_celda=5,
        sumar_pallet_a_bruto=False,  # el total de la plantilla Mexico NO suma el peso de pallets
        total_replacer=lambda fila, tc, tn, tb: _reemplazar_total_celdas(fila, tc, tn, tb),
        armar_nombre_func=armar_nombre_mexico
    )

    # Fechas I.13/I.14/I.15
    f_faena = datos.get('fecha_faena', '')
    f_prod  = datos.get('fecha_produccion', '')
    f_venc  = datos.get('fecha_vencimiento', '')
    trs2 = get_trs(xml)
    xml = _reemplazar_fechas(xml, trs2, f_faena, f_prod, f_venc, fmt_fecha_al)

    # Transporte (buque)
    transporte = datos.get('transporte', '')
    if transporte: xml = xml.replace('>VAPOR/VESSEL:  SUNNY PHOENIX<', '>VAPOR/VESSEL:  ' + transporte + '<')

    # Contenedor
    contenedor = datos.get('contenedor', '')
    if contenedor: xml = xml.replace('>ZMOU8965406<', '>' + contenedor + '<')
    if not contenedor: alertas.append('Contenedor no encontrado - completar manualmente')

    # Precinto combinado AFIP / SENASA
    precinto_afip   = datos.get('precinto_afip', '')
    precinto_senasa = datos.get('precinto_senasa', '')
    if precinto_afip or precinto_senasa:
        combinado = (precinto_afip or '') + ' / ' + (precinto_senasa or '')
        xml = xml.replace('>BAH66487 / 0039365<', '>' + combinado + '<')
    if not precinto_afip:   alertas.append('Precinto AFIP no encontrado - completar manualmente')
    if not precinto_senasa: alertas.append('Precinto SENASA no encontrado - completar manualmente')

    # Temperatura
    es_congelado = datos.get('es_congelado', False)
    xml = _set_temperatura_singapur(xml, es_congelado, tipo_via='maritimo')

    # Fecha de emision
    fecha_emi = datos.get('fecha_emision') or datetime.datetime.now().strftime('%d/%m/%Y')
    xml = xml.replace('>22/04/2026<', '>' + fecha_emi + '<')

    return xml, alertas


# ── TEMPERATURA SINGAPUR ─────────────────────────────────────────────────────

def _set_temperatura_singapur(xml, es_congelado, tipo_via):
    """Todas las plantillas tienen X en refrigeracion por defecto.
    Si es congelado, mover X a congelacion. Si es enfriado, dejar en refrigeracion."""
    if not es_congelado:
        return xml  # X ya esta en refrigeracion, no hacer nada

    # Es congelado: buscar la fila de temperatura y mover la X
    trs = list(re.finditer(r'<w:tr[ >]', xml))
    for i, m in enumerate(trs):
        ini = m.start()
        fin = trs[i+1].start() if i+1 < len(trs) else len(xml)
        fila = xml[ini:fin]
        txt = ''.join(re.findall(r'<w:t[^>]*>([^<]*)</w:t>', fila))
        if 'efrigerac' not in txt or 'ongela' not in txt:
            continue
        if 'X' not in fila:
            continue
        # Fila de temperatura: quitar X de refrigeracion
        nueva_fila = fila.replace('<w:t>X</w:t>', '<w:t></w:t>', 1)
        # Poner X en la celda siguiente a "De congelacion"
        # La celda de congelacion tiene el label, la siguiente es el checkbox (vacio)
        celda_starts = [m2.start() for m2 in re.finditer(r'<w:tc>', nueva_fila)]
        celda_ends   = [m2.start() for m2 in re.finditer(r'</w:tc>', nueva_fila)]
        for idx_c, (cs, ce) in enumerate(zip(celda_starts, celda_ends)):
            bloque = nueva_fila[cs:ce]
            if 'ongela' in ''.join(re.findall(r'<w:t[^>]*>([^<]*)</w:t>', bloque)):
                # La siguiente celda es el checkbox de congelacion
                if idx_c + 1 < len(celda_starts):
                    cs2 = celda_starts[idx_c + 1]
                    ce2 = celda_ends[idx_c + 1]
                    bloque2 = nueva_fila[cs2:ce2]
                    # Si tiene w:t vacio, reemplazarlo; si no, insertar antes de </w:tc>
                    if '<w:t></w:t>' in bloque2:
                        nuevo_bloque2 = bloque2.replace('<w:t></w:t>', '<w:t>X</w:t>', 1)
                    elif re.search(r'<w:t[^>]*></w:t>', bloque2):
                        nuevo_bloque2 = re.sub(r'<w:t[^>]*></w:t>', '<w:t>X</w:t>', bloque2, count=1)
                    else:
                        # Insertar w:r con w:t>X antes de </w:tc>
                        # Buscar el ultimo </w:p> para insertar el run ahi
                        insert_pos = bloque2.rfind('</w:p>')
                        if insert_pos >= 0:
                            nuevo_bloque2 = bloque2[:insert_pos] + '<w:r><w:t>X</w:t></w:r>' + bloque2[insert_pos:]
                        else:
                            nuevo_bloque2 = bloque2
                    nueva_fila = nueva_fila[:cs2] + nuevo_bloque2 + nueva_fila[ce2:]
                break
        xml = xml[:ini] + nueva_fila + xml[fin:]
        break
    return xml


# ── USA W CLASS ───────────────────────────────────────────────────────────────
# Sin desglose producto por producto: un solo total, y la clave es la
# CONTRAMARCA (I.10 Marca de embarque). Fecha de producción y Lote son el
# mismo rango de fechas, solo que en formatos distintos (dd/mm/yyyy vs YYYYMMDD).

def _gen_usa_wclass(xml, datos):
    alertas = []

    total_cajas = str(datos.get('total_cajas', '') or '')
    contramarca = datos.get('contramarca', '') or ''
    if not contramarca:
        alertas.append('Contramarca no encontrada en el remito - completar manualmente')

    f_prod = datos.get('fecha_produccion', '') or ''
    f_venc = datos.get('fecha_vencimiento', '') or ''
    if not f_prod:
        alertas.append('Fecha de producción no encontrada en el piqueo - completar manualmente')
    if not f_venc:
        alertas.append('Fecha límite de conservación no encontrada - completar manualmente')

    peso_neto_kg  = str(datos.get('total_neto', '') or '')
    peso_neto_lbs = kg_a_lbs(peso_neto_kg)

    fecha_prod_fmt = fmt_fecha_al_to_usa(f_prod)
    lote_fmt       = fecha_a_lote_usa(f_prod)
    fecha_venc_fmt = fmt_fecha_al_to_usa(f_venc)
    fecha_emi = datos.get('fecha_emision') or datetime.datetime.now().strftime('%d/%m/%Y')

    # Bultos - misma cifra aparece 3 veces (fila ES, fila EN, fila Totales)
    if total_cajas:
        xml = xml.replace('>239<', '>' + total_cajas + '<')

    # Fecha de producción - aparece 2 veces con valores de ejemplo distintos
    # en la plantilla; ambas se completan con el mismo rango real.
    if fecha_prod_fmt:
        xml = xml.replace('>26/06/2026 al/to 03/07/2026<', '>' + fecha_prod_fmt + '<')
        xml = xml.replace('>30/06/2026 al/to 08/07/2026<', '>' + fecha_prod_fmt + '<')

    # Marca de embarque / Contramarca - aparece 2 veces (fila ES y fila EN)
    if contramarca:
        xml = xml.replace('>C221<', '>' + contramarca + '<')

    # Lote - mismo rango que fecha de producción, en formato YYYYMMDD
    if lote_fmt:
        xml = xml.replace('>20260630 al/to 20260708<', '>' + lote_fmt + '<')

    # Peso neto en KGS y en LBS (el texto va junto a la unidad en el mismo run)
    if peso_neto_kg:
        kg_fmt = peso_neto_kg.replace('.', ',')
        xml = xml.replace('>4594,00 KGS<', '>' + kg_fmt + ' KGS<')
        xml = xml.replace('>4594,00<', '>' + kg_fmt + '<')
    if peso_neto_lbs:
        lbs_fmt = peso_neto_lbs.replace('.', ',')
        xml = xml.replace('>10128,02 LBS<', '>' + lbs_fmt + ' LBS<')
        xml = xml.replace('>10128,02<', '>' + lbs_fmt + '<')

    # Fecha límite de conservación (I.15)
    if fecha_venc_fmt:
        xml = xml.replace('>31/08/2026 al/to 05/11/2026<', '>' + fecha_venc_fmt + '<')

    # Fecha de emisión (pie del certificado)
    xml = xml.replace('>13/07/2026<', '>' + fecha_emi + '<')

    return xml, alertas


# ── USA ORLEANS ───────────────────────────────────────────────────────────────
# Detalle producto por producto en un ANEXO (pagina aparte), 2 filas por producto
# (ES/EN) igual que Wclass. A diferencia de Wclass, cada producto tiene su propia
# Contramarca, Fecha de faena y Fecha de produccion/Lote (via piqueo por Cod Prod
# y provisorio por linea, cruzados por cajas+neto+bruto).

def _gen_usa_orleans(xml, datos):
    alertas = []
    trs = get_trs(xml)

    _, _, _, primera_idx = _get_fila_por_contenido(xml, trs, 'PRODUCTO CRUDO INTACTO')
    if primera_idx is None:
        return xml, ['No se encontro la fila modelo de productos en la plantilla Orleans']

    total_idx = None
    for i in range(primera_idx, len(trs)):
        fila, _, _ = get_fila_xml(xml, trs, i)
        if 'Totales' in fila:
            total_idx = i
            break
    if total_idx is None:
        return xml, ['No se encontro la fila de Totales del anexo en la plantilla Orleans']

    fila_es, ini_mod, _ = get_fila_xml(xml, trs, primera_idx)
    fila_en, _, _       = get_fila_xml(xml, trs, primera_idx + 1)
    ini_totales = trs[total_idx].start()

    nuevas_filas = ''
    for prod in datos.get('productos', []):
        contramarca = prod.get('contramarca', '') or ''
        f_faena = prod.get('fecha_faena_prod', '') or ''
        f_prod  = prod.get('fecha_produccion_prod', '') or ''
        if not contramarca: alertas.append('Producto ' + prod.get('codigo', '') + ': contramarca no encontrada - completar manualmente')
        if not f_faena:      alertas.append('Producto ' + prod.get('codigo', '') + ': fecha de faena no encontrada - completar manualmente')
        if not f_prod:       alertas.append('Producto ' + prod.get('codigo', '') + ': fecha de produccion no encontrada - completar manualmente')

        f_faena_fmt = fmt_fecha_al_to_usa(f_faena)
        f_prod_fmt  = fmt_fecha_al_to_usa(f_prod)
        lote_fmt    = fecha_a_lote_usa(f_prod)
        neto_kg  = formatear_miles(prod.get('neto', '')) + ' KGS'
        neto_lbs = formatear_miles(kg_a_lbs(prod.get('neto', ''))) + ' LBS'

        nueva_es = fila_es
        nueva_es = _reemplazar_celda(nueva_es, 0, str(prod.get('cajas', '')))
        nueva_es = _reemplazar_celda(nueva_es, 1, (prod.get('nombre_es', '') or '').strip().upper())
        nueva_es = _reemplazar_celda(nueva_es, 5, f_faena_fmt)
        nueva_es = _reemplazar_celda(nueva_es, 6, f_prod_fmt)
        nueva_es = _reemplazar_celda(nueva_es, 7, contramarca)
        nueva_es = _reemplazar_celda(nueva_es, 8, lote_fmt)
        nueva_es = _reemplazar_celda(nueva_es, 9, neto_kg)

        nueva_en = fila_en
        nueva_en = _reemplazar_celda(nueva_en, 1, (prod.get('nombre_en', '') or '').strip().upper())
        nueva_en = _reemplazar_celda(nueva_en, 5, f_faena_fmt)
        nueva_en = _reemplazar_celda(nueva_en, 6, f_prod_fmt)
        nueva_en = _reemplazar_celda(nueva_en, 7, contramarca)
        nueva_en = _reemplazar_celda(nueva_en, 8, lote_fmt)
        nueva_en = _reemplazar_celda(nueva_en, 9, neto_lbs)

        nuevas_filas += nueva_es + nueva_en

    xml = xml[:ini_mod] + nuevas_filas + xml[ini_totales:]

    # Totales (aparecen 2 veces: resumen en pagina 1 "VER ANEXO" y al pie del anexo)
    total_cajas = str(datos.get('total_cajas', '') or '')
    total_neto_fmt = formatear_miles(datos.get('total_neto', ''))
    total_lbs_fmt  = formatear_miles(kg_a_lbs(datos.get('total_neto', '')))
    if total_cajas:
        xml = re.sub(r'\b1170\b', total_cajas, xml)
    if total_neto_fmt:
        xml = re.sub(r'21284,00(\s*KGS)', total_neto_fmt + r'\1', xml)
    if total_lbs_fmt:
        xml = re.sub(r'46923,13(\s*LBS)', total_lbs_fmt + r'\1', xml)

    # Transporte (buque - Orleans es maritimo)
    transporte = datos.get('transporte', '') or ''
    if transporte: xml = xml.replace('MAERSK MONTE AZUL', transporte)

    # Contenedor
    contenedor = datos.get('contenedor', '') or ''
    if contenedor: xml = xml.replace('MNBU3586542', contenedor)
    if not contenedor: alertas.append('Contenedor no encontrado - completar manualmente')

    # Precinto (un solo campo en esta plantilla - se usa el de AFIP)
    precinto = datos.get('precinto_afip') or datos.get('precinto_senasa') or ''
    if precinto: xml = xml.replace('BAH74541', precinto)
    if not precinto: alertas.append('Precinto no encontrado - completar manualmente')

    # Fecha limite de conservacion (I.15)
    f_venc_fmt = fmt_fecha_al_to_usa(datos.get('fecha_vencimiento', '') or '')
    if f_venc_fmt: xml = xml.replace('26/08/2026 al/to 30/10/2026', f_venc_fmt)

    # Fecha de emision (aparece 2 veces: certificacion pag.2 y firma del anexo pag.3)
    fecha_emi = datos.get('fecha_emision') or datetime.datetime.now().strftime('%d/%m/%Y')
    xml = xml.replace('21/07/2026', fecha_emi)

    return xml, alertas


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
